from django.shortcuts import render
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from .functions import verify_recaptcha,send_qr_code,generate_tokens,send_verification_email, is_refresh_valid, is_access_valid,decrypt_data,time_left, generate_verification_token
from .models import Students,Coordinators,Subscribers
import re
import hashlib
import jwt
from rest_framework import serializers
from .serializers import StudentSerializer
import threading
import urllib.request
import urllib.parse
import json
# Create your views here.

class TimeLeft(APIView):
    def get(self,request):
        target_date = "08/04/2024 18:30"
        days, hours, minutes, seconds = time_left(target_date)
        if (days <0 or hours <0 or minutes <0 or seconds <0):
            return Response({"days":0,"hours":0,"minutes":0,"seconds":0},status=200)
        return Response({"days":days,"hours":hours,"minutes":minutes,"seconds":seconds},status=200)
class RegisterView(APIView):
    def post(self,request):
        
        try:
            recaptcha_response= request.headers.get('Recaptcha-Token')
            first_name = request.data['first_name']
            last_name = request.data['last_name']
            mobile_number = request.data['mobile_number']
            gender = request.data['gender']
            college_email = request.data['college_email']
            student_id = request.data['student_id']
            branch = request.data['branch']
            section = request.data['section']
            is_hosteler = request.data['is_hosteler']
            hacker_rank_id = request.data['hacker_rank_id']
            isContestOnly = request.data['is_contest_only']
            university_roll_number=request.data['university_roll_number']
            
        except:
             return Response({"message":"Some fields are missing"},status=400)
        #if isContestOnly then
        print(isContestOnly)
        if isContestOnly==True:
            print(isContestOnly)
            pass
        else:
            return Response({"message":"Sorry, Registrations for the workshop is closed now..."},status=400)
        # recaptcha_response="papa"
        user_agent = request.META.get('HTTP_USER_AGENT', '')
        print(user_agent)
        if "(iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15" in user_agent:
            return Response({"message":"Register karle, main sikha dunga"},status=400)
        if "Android 6.0; Nexus 5 Build/MRA58N" in user_agent:
            print(user_agent)
            return Response({"message":"Register karle, main sikha dunga"},status=400)
        url = 'https://www.google.com/recaptcha/api/siteverify'
        values = {
            'secret': settings.RECAPTCHA_SECRET_KEY,
            'response': recaptcha_response
        }
        data = urllib.parse.urlencode(values).encode('utf-8')
        req = urllib.request.Request(url, data)
        response = urllib.request.urlopen(req)
        result = json.load(response)
        print(result)
        try:
            if result['success'] and result['score'] >= 0.5:
                pass
            elif result['success'] and result['score'] == 0.1:
                return Response({"message":"register karle, main sikha dunga"},status=400)
            elif result['success'] and result['score'] == 0.3:
                print(request.data)
                return Response({"message":"bb pr aake register karle, main sikha dunga"},status=400)
            else:
                print(request.data)
                return Response({"message":"Invalid Recaptcha"},status=400)
        except:
            return Response({"message":"Invalid Recaptcha"},status=400)
             
        
        if(settings.PRODUCTION=="TRUE"):
             
            if not re.match(r'^[a-zA-Z0-9_.+-]+@akgec\.ac\.in$', college_email):
                return Response({"message":"Only College email is allowed"},status=400)
            integer_part = re.match(r"\d+", student_id).group(0)
            if integer_part not in college_email:
                return Response({"message":"Invalid Recaptcha"},status=400)
            if not re.match(r"^[a-zA-Z\s'-.]+$", first_name):
                return Response({"message":"First name is invalid"},status=400)
            if not re.match(r"^[a-zA-Z\s'-.]*$", last_name):
                return Response({"message":"Last name is invalid"},status=400)
            if not re.match(r'^[6-9][0-9]{9}$', mobile_number):
                return Response({"message":"Mobile number is invalid"},status=400)
            if not re.match(r'^(22|23)[0-9a-zA-Z_-]+$', student_id):
                return Response({"message": "Only first and second year students are allowed"}, status=400)

            if not re.match(r'^(22|23)\d{11}$', university_roll_number):
                return Response({"message": "Only first and second year students are allowed"}, status=400)

            if hacker_rank_id == "":
                return Response({"message":"Hacker Rank ID can not be blank"},status=400)
            if not isinstance(is_hosteler,bool):
                
                return Response({"message":"is_hosteler must be boolean"},status=400)
            if not isinstance(isContestOnly,bool):
                return Response({"message":"is_ContestOnly must be boolean"},status=400)
            if gender  not in ['Male','Female','Others']:
                return Response({"message":"Invalid gender choice"},status=400)
            
            second_year_students=Students.objects.filter(student_id__startswith='22',isPaid=True,isContestOnly=False).count()
            if second_year_students==30:
                return Response({"message":"Second Year Students Registration Limit Reached"},status=400)


            #print(is_hosteler)
        #if email is in capslock then block
        if college_email.split("@")[0].isupper():
            print("email in caps")
            print(college_email)
            return Response({"message":"Register karle, ye main sikha dunga"},status=400)
        #send_qr_code(college_email,student_id)
        token=generate_verification_token()
        try:
        #check if student already exists and is not verified
            # if(Students.objects.filter(student_id=student_id,isVerified=False).exists()):
            #     print("student exists")
            #     try:
            #         pass
            #     #     student=Students.objects.filter(student_id=student_id).last()
            #     #     student.first_name=first_name
            #     #     student.last_name=last_name
            #     #     student.mobile_number=mobile_number
            #     #     student.gender=gender,
            #     #     student.college_email=college_email,
            #     #     student.student_id=student_id,
            #     #     student.branch=branch,
            #     #     student.section=section,
            #     #     student.isHosteler=is_hosteler,
            #     #     student.hacker_rank_id=hacker_rank_id,
            #     #     student.token=token,
            #     #     student.isContestOnly=isContestOnly,
            #     #     student.token=token
            #     #     student.university_roll_number=university_roll_number
            #     #     student.save()
                
            #     #     email_thread=threading.Thread(
            #     #     target=send_verification_email,
            #     #     args=(college_email,token)
            #     # )
            #     #     email_thread.start()
            #     except Exception as e:
            #         print("error")
            #         print(e)
            #         print(response.data)
            #         return Response({"message":"Something went wrong, Try again later"},status=400)
            if Students.objects.filter(student_id=student_id,isVerified=True).exists():
                return Response({"message":"You have already Registered, Check mail for QR or contact coordinator"},status=400)
            else:
                student = Students(
                    first_name=first_name,
                    last_name=last_name,
                    mobile_number=mobile_number,
                    gender=gender,
                    college_email=college_email,
                    student_id=student_id,
                    branch=branch,
                    section=section,
                    isHosteler=is_hosteler,
                    hacker_rank_id=hacker_rank_id,
                    token=token,
                    isContestOnly=isContestOnly,
                    university_roll_number=university_roll_number
                )
                student.save()
                email_thread=threading.Thread(
                target=send_verification_email,
                args=(college_email,token)
            )
                email_thread.start()
        except Exception as e:
            print(e)
            print(request.data)
            return Response({"message":"Something went wrong, Try again later"},status=400)
        
        
        return Response({'message':'Verification-Email Sent'},status=201)
      
        
        
class LoginView(APIView):
     def post(slef,request):
       
        try:
            username = request.data['username']
            password = request.data['password']

            unique_code=request.data['unique_code']
            
        except:
            return Response({"message":"Some fields are missing"},status=400)
        try:
            coordinator = Coordinators.objects.get(username=username)
        except:
            return Response({"message":"Invalid Credentials"},status=400)
        if coordinator.password==password:
            if coordinator.unique_code=="papa":
                coordinator.unique_code=unique_code
                coordinator.save()
            elif coordinator.unique_code!=unique_code:
                return Response({"message":"Device Unauthorized"},status=400)
            else:
                pass

            access_token, refresh_token = generate_tokens(coordinator.id)
            response_data={
                'access_token':access_token,
                'refresh_token':refresh_token,
                'username':coordinator.username
            }
            return Response(response_data,status=200)
        else:
            return Response({"message":"Invalid Credentials"},status=400)
        
class GetAccessToken(APIView):
    def get(self,request):
        try:
            #get refresh token from query parameter
            refresh_token=request.query_params['refresh_token']
            print(refresh_token)
        except:
            return Response({"message":"No refresh token in body"},status=400)
        if(is_refresh_valid(refresh_token)):
            id=jwt.decode(refresh_token, settings.REFRESH_SECRET_KEY, algorithms=['HS256'])['id']
            access_token, refresh_token = generate_tokens(id)
            return Response({"access_token":access_token},status=200)
        else:
            return Response({"message":"Either the token is invalid or expired"},status=400)



class GetStudentDetails(APIView):
    def get(self,request,std_id):
       
        
        try:
            student=Students.objects.filter(student_id=std_id).last()
            if student is None:
                return Response({"message":"student not found"},status=404)
            serializer = StudentSerializer(student)
            
            return Response(serializer.data,status=200)
        except Exception as e:
            return Response({"message":"student not found"},status=404)

class MakePayment(APIView):
    def get(self,request):
        try:
            token=request.headers['Authorization']
            if(is_access_valid(access_token=token)):
                pass
            else:
                return Response({"message":"Either the token is expired or is invalid"},status=400)
            
        except:
            return Response({"message":"Unauthorized"},status=401)
        
        try:
            qr_data=request.data['qr_data']
            try:
                std_id=decrypt_data(qr_data)
                student=Students.objects.filter(student_id=std_id).last()
                isPaid=student.isPaid
                isContestOnly=student.isContestOnly
                day1_att=student.day1_att
                day2_att=student.day2_att
                contest_att=student.contest_att
                name=student.first_name+" "+student.last_name
                data={
                    "student_id":std_id,
                    "student_name":name,
                    "isPaid":isPaid,
                    "isContestOnly":isContestOnly,
                    "day1_attendance":day1_att,
                    "day2_attendance":day2_att,
                    "contest_attendance":contest_att
                }
               
                return Response(data,status=200)
            except Exception as e:
                return Response({"message":"Invalid QR-Code"},status=400)
        except:
            return Response({"message":"No qr_data in body"},status=400)
    def post(self,request):
        try:
            token=request.headers['Authorization']
            if(is_access_valid(access_token=token)):
                pass
            else:
                return Response({"message":"Either the token is expired or is invalid"},status=400)
            
        except:
            return Response({"message":"Unauthorized"},status=401)
        
        try:
            qr_data=request.data['qr_data']
            try:
                std_id=decrypt_data(qr_data)
                print(std_id)
                student=Students.objects.filter(student_id=std_id).last()
                if(student.isPaid):
                    return Response({"message":"Already Paid"},status=200)
                else:
                    student.isPaid=True
                    student.save()
                    serializer=StudentSerializer(student)
                    return Response({"message":"Payment Successful",
                                     "student_details":serializer.data},status=200)
            except:
                return Response({"message":"Invalid QR Code"},status=400)
        except:
            return Response({"message":"No qr_data in body"},status=400)
from django.shortcuts import redirect      
class VerifyEmail(APIView):
    def get(self,request):
        
        try:
            token=request.query_params['token']
            student=Students.objects.filter(token=token).last()
            print(student)
            if(student.isVerified):
                #redirect to https://programmingclub.tech/already_verified
                return redirect('https://programmingclub.tech/already_verified')

            if(Students.objects.filter(student_id__startswith='22',isPaid=True,isContestOnly=False).count()==30):
                return Response({"message":"30 Students Limit Reached"},status=400)
            student.isVerified=True
            student.save()
            email_thread=threading.Thread(
                target=send_qr_code,
                args=(student.college_email,student.student_id,student.first_name)
            )
            email_thread.start()
           
            return redirect('https://programmingclub.tech/verified')
        except:
            return redirect('https://programmingclub.tech/invalid')
        
class Subscribe(APIView):
    def post(self,request):
        try:
            recaptcha_response= request.headers.get('Recaptcha-Token')
            email=request.data['email']
        except:
            return Response({"message":"Either You are a bot or the email field is blank"},status=400)
        if (verify_recaptcha(recaptcha_response)):
            pass
        else:
            return Response({'message':'Invalid Recaptcha'},status=400)
        if not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            return Response({"message":"Invalid Email"},status=400)
        if(Subscribers.objects.filter(email=email).exists()):
            return Response({"message":"Already Subscribed"},status=400)
        
        subscriber=Subscribers(email=email)
        subscriber.save()
        return Response({"message":"Subscribed Successfully"},status=201)
class Action(APIView):
    def post(self,request):
        try:
            token=request.headers['Authorization']
            if(is_access_valid(access_token=token)):
                pass
            else:
                return Response({"message":"Either the token is expired or is invalid"},status=400)
        except:
            return Response({"message":"Unauthorized"},status=401)
        try:
            qr_data=request.data['qr_data']
            action=request.data['action']
        except:
            return Response({"message":"Some fields are missing"},status=400)
        try:
            std_id=decrypt_data(qr_data)
            student=Students.objects.filter(student_id=std_id).last()
            print(student.token)

            if action=="pay":
                return Response({"message":"Payment Processing Disabled"},status=400)
                if student.student_id.startswith('22'):
                    second_year_students=Students.objects.filter(student_id__startswith='22',isPaid=True,isContestOnly=False).count()
                    if second_year_students==30:
                        return Response({"message":"30 Students Limit Reached"},status=400)
                if(student.isPaid):
                    print(student.isPaid)
                    
                    return Response({"message":"Already Paid"},status=400)
                else:
                    student.isPaid=True
                    student.save()
                    serializer=StudentSerializer(student)
                    return Response({"message":"Payment Successful",
                                     "student_details":serializer.data},status=200)
            elif action=="mark_day1":
                if(student.isPaid==False):
                    return Response({"message":"Payment Pending"},status=400)
                if(student.day1_att):
                    return Response({"message":"Already Marked"},status=400)
                else:
                    student.day1_att=True
                    student.save()
                    serializer=StudentSerializer(student)
                    return Response({"message":"Marked Successfully",
                                     "student_details":serializer.data},status=200)
            elif action=="mark_day2":
                if(student.isPaid==False):
                    return Response({"message":"Payment Pending"},status=400)
                if(student.day2_att):
                    return Response({"message":"Already Marked"},status=400)
                else:
                    student.day2_att=True
                    student.save()
                    serializer=StudentSerializer(student)
                    return Response({"message":"Marked Successfully",
                                     "student_details":serializer.data},status=200)
            elif action=="mark_contest":
                if(student.contest_att):
                    return Response({"message":"Already Marked"},status=400)
                else:
                    student.contest_att=True
                    student.save()
                    serializer=StudentSerializer(student)
                    return Response({"message":"Marked Successfully",
                                     "student_details":serializer.data},status=200)
            else:
                return Response({"message":"Invalid Action"},status=400)
            
            




            
        except:
            return Response({"message":"Invalid QR Code"},status=400)
        

from django.http import FileResponse
from openpyxl import workbook
from openpyxl.utils import get_column_letter
import os       
class GetExcel(APIView):
    def get(self,request,secret):
   
        if secret!=settings.MY_SECRET_KEY:
            return Response({"message":"Invalid Secret Key"},status=400)
        
        if 'workshop' in request.path:
            
            #get students who are not onlycontestants
            students=Students.objects.filter(isContestOnly=False,isVerified=True).order_by('student_id').values('student_id','first_name','mobile_number','isPaid','day1_att','day2_att')

            wb=workbook.Workbook()
            ws=wb.active
            column_names=['Student ID','Name','Mobile','isPaid','Day1_Att','Day2_Att']
            ws.append(['WORKSHOP_ATTENDANCE_SHEET'])  # Add header row
            ws.append([])  # Add empty row before data starts
            ws.append(column_names)  # Add column names row
            for student in students:
                row=[]
                for key in student:
                    if key.startswith('day') or key == 'contest_att':
                        row.append('P' if student[key] else '')
                    elif key == 'isPaid':
                        row.append('Paid' if student[key] else '')

                    else:
                        row.append(student[key])
                ws.append(row)
            
            wb.save('students.xlsx')
            return FileResponse(open('students.xlsx','rb'),as_attachment=True,filename='workshop+contest-attendance_sheet.xlsx')
        elif 'contest' in request.path:
            #get all students 
            students=Students.objects.filter(isVerified=True).order_by('student_id').values('student_id','first_name','mobile_number','contest_att')
            wb=workbook.Workbook()
            ws=wb.active
            column_names=['Student ID','Name','Mobile:','Contest_Att']
            ws.append(['CONTEST_ATTENDANCE_SHEET'])
            ws.append([])
            ws.append(column_names)
            for student in students:
                row=[]
                for key in student:
                    if key == 'contest_att':
                        row.append('P' if student[key] else '')
                    else:
                        row.append(student[key])
                ws.append(row)
            
            wb.save('students.xlsx')
            return FileResponse(open('students.xlsx','rb'),as_attachment=True,filename='contest_attendance_sheet.xlsx')
            
        else:
            return Response({"message":"Invalid URL"},status=400)
        

class VerifyToken(APIView):
    def get(self,request):
        try:
            token=request.headers['Authorization']
        except:
            return Response({"message":"No token in body"},status=400)
        if(is_access_valid(token)):
            return Response({"message":True},status=200)
        else:
            return Response({"message":False},status=400)
        


class GetPaidStudents(APIView):
    def get(self,request,secret):
        if secret!=settings.MY_SECRET_KEY:
            return Response({"message":"Invalid Secret Key"},status=400)
        students=Students.objects.filter(isPaid=True).order_by('student_id').values('student_id','first_name','mobile_number','isPaid')

        wb=workbook.Workbook()
        ws=wb.active
        column_names=['Student ID','Name','Mobile','isPaid']
        ws.append(['PAID_STUDENTS_LIST'])
        ws.append([])
        ws.append(column_names)
        for student in students:
            row=[]
            for key in student:
                if key == 'isPaid':
                    row.append('Paid' if student[key] else '')
                else:
                    row.append(student[key])
            ws.append(row)
        wb.save('students.xlsx')
        return FileResponse(open('students.xlsx','rb'),as_attachment=True,filename='paid_students_list.xlsx')
class GetPaidEmails(APIView):
    def get(self,request,secret):
        if secret != settings.MY_SECRET_KEY:
            return Response({"message":"Invalid Secret Key"},status=400)
        #get paid students emails
        students=Students.objects.filter(isPaid=True).values('first_name','college_email','student_id')
        return Response(students)
        


   
        